#!/usr/bin/env python3
"""
Generate Schedule Excel Template

Creates a properly formatted Excel template with:
- Column headers
- Data validation
- Sample data rows
- Instructions
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path


def create_schedule_template():
    """Create the schedule template Excel file."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule Template"

    # Define column headers
    headers = [
        "Operation",
        "JobID",
        "JobName",
        "SectionID",
        "SectionName",
        "CostCentreID",
        "CostCentreName",
        "StaffID",
        "StaffName",
        "Date",
        "Blocks",
        "StartTime",
        "Notes",
        "IsLocked"
    ]

    # Add headers with styling
    header_fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Add instructions row
    instructions = [
        "CREATE/UPDATE/DELETE",
        "123 (required if no JobName)",
        "Office Renovation (required if no JobID)",
        "1 (optional)",
        "Electrical Work (optional - will prompt if missing)",
        "10 (optional)",
        "Labor (optional - will prompt if missing)",
        "5 (optional)",
        "John Smith (required if no StaffID)",
        "2026-02-12 (required)",
        "4 (required - hours)",
        "09:00 (required - 24hr format, rounded to 15min intervals)",
        "Morning shift (optional)",
        "false (optional)"
    ]

    instructions_fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
    instructions_font = Font(italic=True, size=9, color="6366F1")

    for col_num, instruction in enumerate(instructions, 1):
        cell = ws.cell(row=2, column=col_num, value=instruction)
        cell.fill = instructions_fill
        cell.font = instructions_font
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = border

    # Add sample data rows
    sample_data = [
        ["CREATE", "123", "", "", "Electrical", "", "Labor", "", "John Smith", "2026-02-12", "4", "08:00", "Morning shift", "false"],
        ["CREATE", "", "Warehouse Fix", "", "Plumbing", "", "Materials", "", "Jane Doe", "2026-02-13", "8", "09:00", "Full day", "false"],
        ["UPDATE", "125", "", "2", "", "11", "", "7", "", "2026-02-14", "6", "13:30", "Updated schedule", "false"],
    ]

    for row_num, data_row in enumerate(sample_data, 3):
        for col_num, value in enumerate(data_row, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical='top')

    # Add data validation for Operation column
    operation_validation = DataValidation(
        type="list",
        formula1='"CREATE,UPDATE,DELETE"',
        allow_blank=False
    )
    operation_validation.error = "Please select CREATE, UPDATE, or DELETE"
    operation_validation.errorTitle = "Invalid Operation"
    ws.add_data_validation(operation_validation)
    operation_validation.add(f"A3:A1000")

    # Add data validation for IsLocked column
    boolean_validation = DataValidation(
        type="list",
        formula1='"true,false"',
        allow_blank=True
    )
    ws.add_data_validation(boolean_validation)
    boolean_validation.add(f"N3:N1000")

    # Set column widths
    column_widths = {
        'A': 12,  # Operation
        'B': 10,  # JobID
        'C': 20,  # JobName
        'D': 10,  # SectionID
        'E': 20,  # SectionName
        'F': 12,  # CostCentreID
        'G': 20,  # CostCentreName
        'H': 10,  # StaffID
        'I': 20,  # StaffName
        'J': 12,  # Date
        'K': 8,   # Blocks
        'L': 10,  # StartTime
        'M': 30,  # Notes
        'N': 10,  # IsLocked
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Set row heights
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 60

    # Freeze first two rows
    ws.freeze_panes = "A3"

    # Add a README sheet
    readme_ws = wb.create_sheet("README")
    readme_content = [
        ["Schedule Template Instructions", ""],
        ["", ""],
        ["How to use this template:", ""],
        ["", ""],
        ["1. Fill in schedule data", "Start from row 3 (after the instructions row)"],
        ["", ""],
        ["2. Required columns:", ""],
        ["   • Operation", "Must be CREATE, UPDATE, or DELETE"],
        ["   • JobID OR JobName", "Specify the job (one is required)"],
        ["   • StaffID OR StaffName", "Specify the staff member (one is required)"],
        ["   • Date", "Format: YYYY-MM-DD (e.g., 2026-02-12)"],
        ["   • Blocks", "Number of hours (integer)"],
        ["", ""],
        ["3. Optional columns:", ""],
        ["   • SectionID/SectionName", "If not provided, you'll be prompted to choose"],
        ["   • CostCentreID/CostCentreName", "If not provided, you'll be prompted to choose"],
        ["   • Notes", "Any additional information"],
        ["   • IsLocked", "true or false (default: false)"],
        ["", ""],
        ["4. Using IDs vs Names:", ""],
        ["   • IDs are faster and more accurate", "Use if you know them"],
        ["   • Names are user-friendly", "System will resolve them automatically"],
        ["   • Mix and match", "You can use IDs for some rows and names for others"],
        ["", ""],
        ["5. For UPDATE operations:", ""],
        ["   • ScheduleID is required", "Add a 'ScheduleID' column if updating"],
        ["   • All other fields optional", "Only include fields you want to change"],
        ["", ""],
        ["6. For DELETE operations:", ""],
        ["   • ScheduleID is required", "Add a 'ScheduleID' column"],
        ["   • Path IDs required", "JobID/QuoteID, SectionID, CostCentreID needed for API path"],
        ["   • Other fields ignored", "You can leave them empty"],
        ["", ""],
        ["7. Upload to Optificial:", ""],
        ["   • Save this file", "Keep the .xlsx format"],
        ["   • Upload in chat", "Use the attachment button"],
        ["   • Include keywords", "Type 'create schedules' or 'bulk schedule create'"],
        ["", ""],
        ["8. Handling errors:", ""],
        ["   • ≤5 errors", "You'll see an interactive form to fix them"],
        ["   • >5 errors", "Download a corrected template with issues highlighted"],
        ["", ""],
        ["Need help?", "Contact support or check the documentation"],
    ]

    for row_num, (col1, col2) in enumerate(readme_content, 1):
        readme_ws.cell(row=row_num, column=1, value=col1)
        readme_ws.cell(row=row_num, column=2, value=col2)

        # Style title row
        if row_num == 1:
            readme_ws.cell(row=row_num, column=1).font = Font(bold=True, size=14, color="8B5CF6")
        # Style section headers
        elif col1 and col1.endswith(":"):
            readme_ws.cell(row=row_num, column=1).font = Font(bold=True, size=11)

    readme_ws.column_dimensions['A'].width = 35
    readme_ws.column_dimensions['B'].width = 60

    return wb


def main():
    """Generate and save the template."""
    template_dir = Path(__file__).parent.parent / "static" / "templates"
    template_dir.mkdir(parents=True, exist_ok=True)

    template_path = template_dir / "schedule_template.xlsx"

    print(f"Generating schedule template at: {template_path}")
    wb = create_schedule_template()
    wb.save(template_path)
    print(f"[OK] Template created successfully!")


if __name__ == "__main__":
    main()
