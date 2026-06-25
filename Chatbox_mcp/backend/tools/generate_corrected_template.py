"""
Generate a pre-filled corrected Excel template for schedule clarifications.

Creates an Excel file with:
- User's original data pre-filled
- Issue cells highlighted (yellow for ambiguous, red for missing)
- Dropdown lists for ambiguous matches (like the clarification form)
- Comment annotations explaining the issue
- Successfully resolved rows included as-is
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

logger = logging.getLogger(__name__)

# ── Style constants ──────────────────────────────────────────────────────
_HEADER_FILL = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
_AMBIGUOUS_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # yellow
_MISSING_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")  # red
_OK_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")  # green

# Map from field name in clarification to the ID + Name column pair.
# When user selects from a dropdown (Name column), we also need to fill the ID column.
_FIELD_TO_ID_COL: Dict[str, str] = {
    "StaffName": "StaffID",
    "JobName": "JobID",
    "SectionName": "SectionID",
    "CostCentreName": "CostCentreID",
    "SiteName": "SiteName",
}


def generate_corrected_template(
    extracted: Dict[str, Any],
    clarifications: List[Dict[str, Any]],
    session_id: str,
    output_dir: Optional[Path] = None,
) -> Path:
    """
    Generate a corrected Excel file pre-filled with user data and issue markers.

    Args:
        extracted: Original extracted data {"tables": [{"headers": [...], "rows": [...]}]}
        clarifications: List of clarification dicts from the schedule agent
        session_id: Session ID for the filename
        output_dir: Directory to save the file (defaults to static/temp/)

    Returns:
        Path to the generated Excel file
    """
    if output_dir is None:
        output_dir = Path(__file__).parent.parent / "static" / "temp"
    output_dir.mkdir(parents=True, exist_ok=True)

    table = extracted.get("tables", [{}])[0]
    headers = table.get("headers", [])
    rows = table.get("rows", [])

    # Build a lookup: (row_index, field_name) → clarification
    # Row indices in clarifications are 2-based (row 2 = first data row)
    clar_map: Dict[tuple, Dict[str, Any]] = {}
    for clar in clarifications:
        key = (clar["row"], clar["field"])
        clar_map[key] = clar

    wb = Workbook()
    ws = wb.active
    ws.title = "Corrected Schedule"

    # ── Row 1: Headers ──
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER

    # ── Data rows ──
    # Build header→column index map
    header_col: Dict[str, int] = {h: i + 1 for i, h in enumerate(headers)}

    for row_idx, row_data in enumerate(rows):
        excel_row = row_idx + 2  # Excel row (1-based, row 1 is headers)
        agent_row = row_idx + 2  # Agent row numbering (2-based)

        # row_data can be a list (positional) or dict
        if isinstance(row_data, list):
            for col_idx, value in enumerate(row_data):
                if col_idx < len(headers):
                    cell = ws.cell(row=excel_row, column=col_idx + 1, value=str(value) if value else "")
                    cell.border = _BORDER
                    cell.alignment = Alignment(vertical="top")
        elif isinstance(row_data, dict):
            for header in headers:
                col_idx = header_col[header]
                value = row_data.get(header, "")
                cell = ws.cell(row=excel_row, column=col_idx, value=str(value) if value else "")
                cell.border = _BORDER
                cell.alignment = Alignment(vertical="top")

    # ── Apply clarification markers ──
    # Track which rows have issues
    issue_rows = set()

    for (agent_row, field), clar in clar_map.items():
        excel_row = agent_row  # agent_row is already 2-based matching excel rows
        col_idx = header_col.get(field)
        if col_idx is None:
            continue

        issue_rows.add(excel_row)
        cell = ws.cell(row=excel_row, column=col_idx)
        clar_type = clar.get("type", "")

        if clar_type == "ambiguous" and clar.get("options"):
            # Yellow highlight + dropdown with options
            cell.fill = _AMBIGUOUS_FILL

            options = clar["options"]
            # Build dropdown formula: "option1,option2,option3"
            # For staff/job/etc, show "Name (ID: xxx)" format for clarity
            dropdown_items = []
            for opt in options:
                opt_name = opt.get("name", "")
                opt_id = opt.get("id", "")
                if opt_id:
                    dropdown_items.append(f"{opt_name} (ID:{opt_id})")
                else:
                    dropdown_items.append(opt_name)

            # Excel data validation has a 255-char limit for inline formulas.
            # If options are too long, truncate to fit.
            formula_str = ",".join(dropdown_items)
            if len(formula_str) > 255:
                # Trim options to fit
                trimmed = []
                length = 0
                for item in dropdown_items:
                    if length + len(item) + 1 > 255:
                        break
                    trimmed.append(item)
                    length += len(item) + 1
                formula_str = ",".join(trimmed)

            if formula_str:
                dv = DataValidation(
                    type="list",
                    formula1=f'"{formula_str}"',
                    allow_blank=True,
                )
                dv.error = f"Please select a valid {field}"
                dv.errorTitle = f"Ambiguous {field}"
                dv.prompt = f"Multiple matches found for '{clar.get('value', '')}'. Select the correct one."
                dv.promptTitle = "Select match"
                dv.showInputMessage = True
                ws.add_data_validation(dv)
                cell_ref = ws.cell(row=excel_row, column=col_idx).coordinate
                dv.add(cell_ref)

            # Add comment with full context
            comment_text = f"AMBIGUOUS: '{clar.get('value', '')}' matched multiple entries.\n"
            comment_text += "Select from the dropdown or type the correct value.\n\n"
            comment_text += "Options:\n"
            for opt in options[:10]:
                comment_text += f"  • {opt.get('name', '')} (ID: {opt.get('id', '')})\n"
            cell.comment = Comment(comment_text, "Optificial AI", width=300, height=200)

        elif clar_type in ("missing", "free_text"):
            # Red highlight for missing fields
            cell.fill = _MISSING_FILL

            comment_text = f"MISSING: {clar.get('message', field + ' is required')}\n"
            if clar.get("options"):
                comment_text += "\nAvailable options:\n"
                for opt in clar["options"][:10]:
                    if isinstance(opt, dict):
                        comment_text += f"  • {opt.get('name', opt.get('Name', ''))}\n"
                    else:
                        comment_text += f"  • {opt}\n"
                # Also add dropdown for missing fields with options
                dropdown_items = []
                for opt in clar["options"]:
                    if isinstance(opt, dict):
                        opt_name = opt.get("name", opt.get("Name", ""))
                        opt_id = opt.get("id", opt.get("ID", ""))
                        if opt_id:
                            dropdown_items.append(f"{opt_name} (ID:{opt_id})")
                        else:
                            dropdown_items.append(opt_name)
                    else:
                        dropdown_items.append(str(opt))

                formula_str = ",".join(dropdown_items)
                if len(formula_str) > 255:
                    trimmed = []
                    length = 0
                    for item in dropdown_items:
                        if length + len(item) + 1 > 255:
                            break
                        trimmed.append(item)
                        length += len(item) + 1
                    formula_str = ",".join(trimmed)

                if formula_str:
                    dv = DataValidation(
                        type="list",
                        formula1=f'"{formula_str}"',
                        allow_blank=True,
                    )
                    dv.error = f"Please select a valid {field}"
                    dv.prompt = clar.get("message", f"Select or enter {field}")
                    dv.showInputMessage = True
                    ws.add_data_validation(dv)
                    cell_ref = ws.cell(row=excel_row, column=col_idx).coordinate
                    dv.add(cell_ref)
            elif clar.get("placeholder"):
                comment_text += f"\nExpected format: {clar['placeholder']}"

            cell.comment = Comment(comment_text, "Optificial AI", width=300, height=150)

    # ── Mark OK rows with green tint ──
    for row_idx in range(len(rows)):
        excel_row = row_idx + 2
        if excel_row not in issue_rows:
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=excel_row, column=col_idx)
                if not cell.fill or cell.fill.start_color.index == "00000000":
                    cell.fill = _OK_FILL

    # ── Column widths (auto-fit based on content) ──
    for col_idx, header in enumerate(headers, 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        max_len = len(header)
        for row_idx in range(2, len(rows) + 2):
            val = str(ws.cell(row=row_idx, column=col_idx).value or "")
            max_len = max(max_len, min(len(val), 40))
        ws.column_dimensions[col_letter].width = max_len + 4

    # Freeze header row
    ws.freeze_panes = "A2"

    # ── Add Issues Summary sheet ──
    summary_ws = wb.create_sheet("Issues Summary")
    summary_ws.cell(row=1, column=1, value="Row").font = Font(bold=True)
    summary_ws.cell(row=1, column=2, value="Field").font = Font(bold=True)
    summary_ws.cell(row=1, column=3, value="Issue Type").font = Font(bold=True)
    summary_ws.cell(row=1, column=4, value="Current Value").font = Font(bold=True)
    summary_ws.cell(row=1, column=5, value="Message").font = Font(bold=True)

    for i, clar in enumerate(clarifications, 2):
        summary_ws.cell(row=i, column=1, value=clar.get("row", ""))
        summary_ws.cell(row=i, column=2, value=clar.get("field", ""))
        summary_ws.cell(row=i, column=3, value=clar.get("type", "").upper())
        summary_ws.cell(row=i, column=4, value=clar.get("value", ""))
        summary_ws.cell(row=i, column=5, value=clar.get("message", ""))

    summary_ws.column_dimensions["A"].width = 8
    summary_ws.column_dimensions["B"].width = 18
    summary_ws.column_dimensions["C"].width = 18
    summary_ws.column_dimensions["D"].width = 25
    summary_ws.column_dimensions["E"].width = 50

    # Save
    output_path = output_dir / f"corrected_{session_id}.xlsx"
    wb.save(output_path)
    logger.info(f"📄 Generated corrected template: {output_path} ({len(clarifications)} issues, {len(rows)} rows)")

    return output_path
